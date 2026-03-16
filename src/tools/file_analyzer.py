"""
File Upload Analysis (OpenAI Parity)
=====================================
Analyzes uploaded files: Excel, CSV, PDF, etc.

Supports:
- Excel formula analysis
- CSV data analysis
- PDF text + table extraction
- Statistical analysis
- Chart generation
"""

import logging
from typing import Any, Dict, List, Optional
from pathlib import Path
from dataclasses import dataclass

logger = logging.getLogger(__name__)


@dataclass
class FileAnalysisResult:
    """Result from file analysis."""
    file_type: str
    file_name: str
    summary: str
    data: Any
    statistics: Dict[str, Any]
    insights: List[str]
    charts: List[Dict]  # Chart specifications


class FileAnalyzer:
    """
    Comprehensive file analyzer.

    Handles multiple file types with deep analysis.
    """

    def __init__(
        self,
        auto_generate_charts: bool = False,
        chart_output_dir: str = "outputs/charts",
    ):
        """
        Initialize file analyzer.

        Args:
            auto_generate_charts: If True, automatically generate chart images
            chart_output_dir: Directory for generated chart images
        """
        self.auto_generate_charts = auto_generate_charts
        self.chart_output_dir = chart_output_dir
        self._chart_generator = None
        self._check_dependencies()

    def _check_dependencies(self):
        """Check available analysis libraries."""
        self.has_pandas = False
        self.has_openpyxl = False
        self.has_pdfplumber = False

        try:
            import pandas
            self.has_pandas = True
        except ImportError:
            logger.warning("[FILE] pandas not available")

        try:
            import openpyxl
            self.has_openpyxl = True
        except ImportError:
            logger.warning("[FILE] openpyxl not available")

        try:
            import pdfplumber
            self.has_pdfplumber = True
        except ImportError:
            logger.warning("[FILE] pdfplumber not available")

    @property
    def chart_generator(self):
        """Lazy-load chart generator."""
        if self._chart_generator is None:
            self._chart_generator = ChartGenerator(output_dir=self.chart_output_dir)
        return self._chart_generator

    def analyze_file(
        self,
        file_path: str,
        generate_charts: Optional[bool] = None,
    ) -> FileAnalysisResult:
        """
        Analyze any supported file type.

        Args:
            file_path: Path to file to analyze
            generate_charts: Override auto_generate_charts setting for this call
        """
        path = Path(file_path)
        suffix = path.suffix.lower()

        if suffix in ['.xlsx', '.xls']:
            result = self.analyze_excel(file_path)
        elif suffix == '.csv':
            result = self.analyze_csv(file_path)
        elif suffix == '.pdf':
            result = self.analyze_pdf(file_path)
        elif suffix == '.json':
            result = self.analyze_json(file_path)
        else:
            return FileAnalysisResult(
                file_type="unknown",
                file_name=path.name,
                summary="Unsupported file type",
                data=None,
                statistics={},
                insights=[],
                charts=[],
            )

        # Auto-generate charts if enabled
        should_generate = generate_charts if generate_charts is not None else self.auto_generate_charts

        if should_generate and result.charts:
            generated_paths = self.chart_generator.generate_from_suggestions(
                analysis_result=result,
                max_charts=5,
            )

            if generated_paths:
                logger.info(f"[FILE] Generated {len(generated_paths)} charts: {generated_paths}")
                # Update the result with generated chart paths
                result.charts = [
                    {**chart, "generated_path": path}
                    for chart, path in zip(result.charts, generated_paths)
                ]

        return result

    def analyze_excel(self, file_path: str) -> FileAnalysisResult:
        """Analyze Excel file with formula detection."""
        if not self.has_pandas or not self.has_openpyxl:
            return self._error_result("Excel", file_path, "Missing pandas/openpyxl")

        import pandas as pd
        import openpyxl

        path = Path(file_path)

        # Load with openpyxl for formulas
        wb = openpyxl.load_workbook(file_path, data_only=False)

        sheets_data = {}
        formulas = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Extract formulas
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).startswith('='):
                        formulas.append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "formula": cell.value,
                        })

        # Load with pandas for data analysis
        df_dict = pd.read_excel(file_path, sheet_name=None)

        all_stats = {}
        insights = []

        for sheet_name, df in df_dict.items():
            sheets_data[sheet_name] = df.to_dict('records')[:100]  # First 100 rows

            # Calculate statistics
            numeric_cols = df.select_dtypes(include=['number']).columns
            if len(numeric_cols) > 0:
                stats = df[numeric_cols].describe().to_dict()
                all_stats[sheet_name] = stats

                # Generate insights
                for col in numeric_cols:
                    insights.append(f"{sheet_name}.{col}: mean={df[col].mean():.2f}, std={df[col].std():.2f}")

        return FileAnalysisResult(
            file_type="excel",
            file_name=path.name,
            summary=f"Excel file with {len(wb.sheetnames)} sheets, {len(formulas)} formulas",
            data={"sheets": sheets_data, "formulas": formulas},
            statistics=all_stats,
            insights=insights,
            charts=self._suggest_charts(df_dict),
        )

    def analyze_csv(self, file_path: str) -> FileAnalysisResult:
        """Analyze CSV file."""
        if not self.has_pandas:
            return self._error_result("CSV", file_path, "Missing pandas")

        import pandas as pd

        path = Path(file_path)
        df = pd.read_csv(file_path)

        # Basic statistics
        numeric_cols = df.select_dtypes(include=['number']).columns
        stats = df[numeric_cols].describe().to_dict() if len(numeric_cols) > 0 else {}

        # Generate insights
        insights = []
        insights.append(f"Shape: {df.shape[0]} rows, {df.shape[1]} columns")

        for col in numeric_cols:
            insights.append(f"{col}: range=[{df[col].min():.2f}, {df[col].max():.2f}]")

        # Check for missing values
        missing = df.isnull().sum()
        if missing.any():
            insights.append(f"Missing values in: {list(missing[missing > 0].index)}")

        return FileAnalysisResult(
            file_type="csv",
            file_name=path.name,
            summary=f"CSV file with {df.shape[0]} rows, {df.shape[1]} columns",
            data=df.head(100).to_dict('records'),
            statistics=stats,
            insights=insights,
            charts=self._suggest_charts({"data": df}),
        )

    def analyze_pdf(self, file_path: str) -> FileAnalysisResult:
        """Analyze PDF file."""
        if not self.has_pdfplumber:
            return self._error_result("PDF", file_path, "Missing pdfplumber")

        import pdfplumber

        path = Path(file_path)

        text_content = []
        tables = []

        with pdfplumber.open(file_path) as pdf:
            for i, page in enumerate(pdf.pages[:20]):  # First 20 pages
                # Extract text
                text = page.extract_text()
                if text:
                    text_content.append({"page": i+1, "text": text})

                # Extract tables
                page_tables = page.extract_tables()
                for j, table in enumerate(page_tables):
                    tables.append({
                        "page": i+1,
                        "table_index": j,
                        "data": table,
                    })

        # Generate insights
        insights = []
        insights.append(f"Total pages processed: {min(len(text_content), 20)}")
        insights.append(f"Tables found: {len(tables)}")

        total_words = sum(len(p["text"].split()) for p in text_content)
        insights.append(f"Total words: {total_words}")

        return FileAnalysisResult(
            file_type="pdf",
            file_name=path.name,
            summary=f"PDF with {len(text_content)} pages, {len(tables)} tables",
            data={"text": text_content, "tables": tables},
            statistics={"pages": len(text_content), "tables": len(tables), "words": total_words},
            insights=insights,
            charts=[],
        )

    def analyze_json(self, file_path: str) -> FileAnalysisResult:
        """Analyze JSON file."""
        import json

        path = Path(file_path)

        with open(file_path, 'r') as f:
            data = json.load(f)

        # Analyze structure
        def analyze_structure(obj, path=""):
            if isinstance(obj, dict):
                return {k: analyze_structure(v, f"{path}.{k}") for k, v in list(obj.items())[:10]}
            elif isinstance(obj, list):
                return f"list[{len(obj)}]"
            else:
                return type(obj).__name__

        structure = analyze_structure(data)

        return FileAnalysisResult(
            file_type="json",
            file_name=path.name,
            summary="JSON file",
            data=data if isinstance(data, list) and len(data) < 100 else {"sample": str(data)[:1000]},
            statistics={"structure": structure},
            insights=[f"Top-level type: {type(data).__name__}"],
            charts=[],
        )

    def _suggest_charts(self, df_dict: Dict) -> List[Dict]:
        """Suggest appropriate charts for the data."""
        suggestions = []

        for sheet_name, df in df_dict.items():
            if hasattr(df, 'select_dtypes'):
                numeric = df.select_dtypes(include=['number']).columns
                categorical = df.select_dtypes(include=['object']).columns

                if len(numeric) >= 2:
                    suggestions.append({
                        "type": "scatter",
                        "x": numeric[0],
                        "y": numeric[1],
                        "sheet": sheet_name,
                    })

                if len(categorical) > 0 and len(numeric) > 0:
                    suggestions.append({
                        "type": "bar",
                        "x": categorical[0],
                        "y": numeric[0],
                        "sheet": sheet_name,
                    })

        return suggestions

    def _error_result(self, file_type: str, file_path: str, error: str) -> FileAnalysisResult:
        """Return error result."""
        return FileAnalysisResult(
            file_type=file_type,
            file_name=Path(file_path).name,
            summary=f"Error: {error}",
            data=None,
            statistics={},
            insights=[error],
            charts=[],
        )


class ChartGenerator:
    """
    Generates charts from data analysis results.

    Supports:
    - Bar charts
    - Line charts
    - Scatter plots
    - Pie charts
    - Histograms
    - Heatmaps
    """

    def __init__(self, output_dir: str = "outputs/charts"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self._check_matplotlib()

    def _check_matplotlib(self):
        """Check if matplotlib is available."""
        self.has_matplotlib = False
        self.has_seaborn = False

        try:
            import matplotlib
            matplotlib.use('Agg')  # Non-interactive backend
            import matplotlib.pyplot as plt
            self.has_matplotlib = True
        except ImportError:
            logger.warning("[CHART] matplotlib not available")

        try:
            import seaborn
            self.has_seaborn = True
        except ImportError:
            logger.warning("[CHART] seaborn not available (optional)")

    def generate_chart(
        self,
        data: Any,
        chart_type: str,
        x_column: Optional[str] = None,
        y_column: Optional[str] = None,
        title: Optional[str] = None,
        filename: Optional[str] = None,
    ) -> Optional[str]:
        """
        Generate a chart from data.

        Args:
            data: DataFrame or dict with data
            chart_type: Type of chart (bar, line, scatter, pie, histogram, heatmap)
            x_column: Column for x-axis
            y_column: Column for y-axis
            title: Chart title
            filename: Output filename (auto-generated if not provided)

        Returns:
            Path to generated chart image, or None if failed
        """
        if not self.has_matplotlib:
            logger.error("[CHART] Cannot generate chart: matplotlib not installed")
            return None

        import matplotlib.pyplot as plt

        # Convert to DataFrame if needed
        if not self.has_pandas:
            try:
                import pandas as pd
                self.has_pandas = True
            except ImportError:
                logger.error("[CHART] pandas required for chart generation")
                return None
        else:
            import pandas as pd

        if isinstance(data, dict):
            df = pd.DataFrame(data)
        elif hasattr(data, 'to_frame'):
            df = data.to_frame()
        else:
            df = data

        # Create figure
        fig, ax = plt.subplots(figsize=(10, 6))

        try:
            if chart_type == "bar":
                self._create_bar_chart(ax, df, x_column, y_column)
            elif chart_type == "line":
                self._create_line_chart(ax, df, x_column, y_column)
            elif chart_type == "scatter":
                self._create_scatter_chart(ax, df, x_column, y_column)
            elif chart_type == "pie":
                self._create_pie_chart(ax, df, x_column, y_column)
            elif chart_type == "histogram":
                self._create_histogram(ax, df, x_column)
            elif chart_type == "heatmap":
                self._create_heatmap(fig, ax, df)
            else:
                logger.warning(f"[CHART] Unsupported chart type: {chart_type}")
                plt.close(fig)
                return None

            # Set title
            if title:
                ax.set_title(title, fontsize=14, fontweight='bold')
            else:
                ax.set_title(f"{chart_type.title()} Chart", fontsize=14)

            # Generate filename
            if not filename:
                from datetime import datetime
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"chart_{chart_type}_{timestamp}.png"

            output_path = self.output_dir / filename

            # Save chart
            plt.tight_layout()
            plt.savefig(output_path, dpi=150, bbox_inches='tight')
            plt.close(fig)

            logger.info(f"[CHART] Generated: {output_path}")
            return str(output_path)

        except Exception as e:
            logger.error(f"[CHART] Failed to generate {chart_type} chart: {e}")
            plt.close(fig)
            return None

    def _create_bar_chart(self, ax, df, x_col, y_col):
        """Create a bar chart."""
        if x_col and y_col:
            df.plot(kind='bar', x=x_col, y=y_col, ax=ax, legend=False)
            ax.set_xlabel(x_col)
            ax.set_ylabel(y_col)
        else:
            # Use first categorical and first numeric
            numeric_cols = df.select_dtypes(include=['number']).columns
            if len(numeric_cols) > 0:
                df[numeric_cols[0]].head(20).plot(kind='bar', ax=ax)
                ax.set_ylabel(numeric_cols[0])

        ax.tick_params(axis='x', rotation=45)

    def _create_line_chart(self, ax, df, x_col, y_col):
        """Create a line chart."""
        if x_col and y_col:
            df.plot(kind='line', x=x_col, y=y_col, ax=ax, marker='o')
            ax.set_xlabel(x_col)
            ax.set_ylabel(y_col)
        else:
            numeric_cols = df.select_dtypes(include=['number']).columns
            if len(numeric_cols) > 0:
                df[numeric_cols].plot(kind='line', ax=ax, marker='o')

        ax.legend(loc='best')
        ax.grid(True, alpha=0.3)

    def _create_scatter_chart(self, ax, df, x_col, y_col):
        """Create a scatter plot."""
        if x_col and y_col:
            ax.scatter(df[x_col], df[y_col], alpha=0.6, edgecolors='w', linewidth=0.5)
            ax.set_xlabel(x_col)
            ax.set_ylabel(y_col)
        else:
            numeric_cols = df.select_dtypes(include=['number']).columns
            if len(numeric_cols) >= 2:
                ax.scatter(df[numeric_cols[0]], df[numeric_cols[1]], alpha=0.6)
                ax.set_xlabel(numeric_cols[0])
                ax.set_ylabel(numeric_cols[1])

        ax.grid(True, alpha=0.3)

    def _create_pie_chart(self, ax, df, x_col, y_col):
        """Create a pie chart."""
        if y_col:
            values = df[y_col].head(10)  # Limit to 10 slices
            labels = df[x_col].head(10) if x_col else values.index
        else:
            numeric_cols = df.select_dtypes(include=['number']).columns
            if len(numeric_cols) > 0:
                values = df[numeric_cols[0]].head(10)
                labels = df.index[:10]
            else:
                return

        ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90)
        ax.axis('equal')

    def _create_histogram(self, ax, df, x_col):
        """Create a histogram."""
        if x_col:
            df[x_col].hist(ax=ax, bins=30, edgecolor='white', alpha=0.7)
            ax.set_xlabel(x_col)
        else:
            numeric_cols = df.select_dtypes(include=['number']).columns
            if len(numeric_cols) > 0:
                df[numeric_cols[0]].hist(ax=ax, bins=30, edgecolor='white', alpha=0.7)
                ax.set_xlabel(numeric_cols[0])

        ax.set_ylabel('Frequency')
        ax.grid(True, alpha=0.3)

    def _create_heatmap(self, fig, ax, df):
        """Create a heatmap of correlations."""
        numeric_df = df.select_dtypes(include=['number'])

        if numeric_df.shape[1] < 2:
            logger.warning("[CHART] Need at least 2 numeric columns for heatmap")
            return

        corr_matrix = numeric_df.corr()

        if self.has_seaborn:
            import seaborn as sns
            sns.heatmap(
                corr_matrix,
                ax=ax,
                annot=True,
                cmap='coolwarm',
                center=0,
                fmt='.2f',
                square=True,
            )
        else:
            im = ax.imshow(corr_matrix.values, cmap='coolwarm', aspect='auto')
            ax.set_xticks(range(len(corr_matrix.columns)))
            ax.set_yticks(range(len(corr_matrix.columns)))
            ax.set_xticklabels(corr_matrix.columns, rotation=45, ha='right')
            ax.set_yticklabels(corr_matrix.columns)
            fig.colorbar(im, ax=ax)

    def generate_from_suggestions(
        self,
        analysis_result: FileAnalysisResult,
        max_charts: int = 5,
    ) -> List[str]:
        """
        Generate charts based on analysis suggestions.

        Args:
            analysis_result: Result from FileAnalyzer
            max_charts: Maximum number of charts to generate

        Returns:
            List of paths to generated chart images
        """
        chart_paths = []
        suggestions = analysis_result.charts[:max_charts]

        for i, suggestion in enumerate(suggestions):
            chart_type = suggestion.get("type", "bar")
            x_col = suggestion.get("x")
            y_col = suggestion.get("y")
            sheet = suggestion.get("sheet", "data")

            # Get data for this chart
            data = analysis_result.data
            if isinstance(data, dict) and "sheets" in data:
                data = data["sheets"].get(sheet, {})

            if not data:
                continue

            path = self.generate_chart(
                data=data,
                chart_type=chart_type,
                x_column=x_col,
                y_column=y_col,
                title=f"{sheet}: {y_col or 'values'} by {x_col or 'index'}",
                filename=f"chart_{i+1}_{chart_type}.png",
            )

            if path:
                chart_paths.append(path)

        return chart_paths

    @property
    def has_pandas(self):
        """Check if pandas is available."""
        try:
            import pandas
            return True
        except ImportError:
            return False
